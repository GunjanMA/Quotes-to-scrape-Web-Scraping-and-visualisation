from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep
from parsel import Selector
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt

driver=webdriver.Chrome('C:/Users/gunjan/Desktop/Web_Scraping/chromedriver')
driver.get('http://quotes.toscrape.com/')


sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()

#next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
#next_btn.click()

path="C:/Users/gunjan/Desktop/quotes_web_scrap/data.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.active

for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+1,column=1).value=quotes[r-1]
        sheet.cell(row=r+1,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+1,column=3).value=about_links[r-1]
            #i=i+1
next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+10,column=1).value=quotes[r-1]
        sheet.cell(row=r+10,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+10,column=3).value=about_links[r-1]
            #i=i+1
next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+20,column=1).value=quotes[r-1]
        sheet.cell(row=r+20,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+20,column=3).value=about_links[r-1]
            #i=i+1

next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+30,column=1).value=quotes[r-1]
        sheet.cell(row=r+30,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+30,column=3).value=about_links[r-1]
            #i=i+1

next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+40,column=1).value=quotes[r-1]
        sheet.cell(row=r+40,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+40,column=3).value=about_links[r-1]
            #i=i+1

next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+50,column=1).value=quotes[r-1]
        sheet.cell(row=r+50,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+50,column=3).value=about_links[r-1]
            #i=i+1

next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+60,column=1).value=quotes[r-1]
        sheet.cell(row=r+60,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+60,column=3).value=about_links[r-1]
            #i=i+1
next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+70,column=1).value=quotes[r-1]
        sheet.cell(row=r+70,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+70,column=3).value=about_links[r-1]
            #i=i+1
next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+80,column=1).value=quotes[r-1]
        sheet.cell(row=r+80,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+80,column=3).value=about_links[r-1]
            #i=i+1
next_btn=driver.find_element_by_xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "next", " " ))]//a')
next_btn.click()

sel=Selector(text=driver.page_source)
quotes=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "text", " " ))]/text()').extract()
author=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "author", " " ))]/text()').extract()
#tags=sel.xpath('//*[contains(concat( " ", @class, " " ), concat( " ", "tag", " " ))]/text()').extract()
about_links=sel.xpath('//span//a').extract()


for r in range(1,len(quotes)):
        #for c in range(1,4):
        sheet.cell(row=r+90,column=1).value=quotes[r-1]
        sheet.cell(row=r+90,column=2).value=author[r-1]
        #sheet.cell(row=r,column=3).value=tags[r-1]
        sheet.cell(row=r+90,column=3).value=about_links[r-1]
            #i=i+1
        

workbook.save(path)

d=pd.read_excel('C:\\Users\\gunjan\\Desktop\\quotes_web_scrap\\data.xlsx')

plt.plot(d.groupby('Author').Author.count())
plt.xticks(rotation='vertical')
plt.show()
